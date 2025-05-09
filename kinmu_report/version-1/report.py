import os
import openpyxl
import sys
import tkinter as tk
import re
from datetime import datetime
from tkinter import messagebox, filedialog

# Precompile regex patterns for normalization
TIME_PATTERN = re.compile(r'(\d{1,3}:\d{2})\s+\(')
CLEAN_NAME_PATTERN = re.compile(r'^[\d._\-]+')

def create_root():
    root = tk.Tk()
    root.withdraw()
    return root

def select_directory(root, prompt):
    messagebox.showinfo('情報', prompt)
    return filedialog.askdirectory(title=prompt)

def get_comparison_columns(col, sheet_name):
    if "勤務表" in sheet_name:
        if col < 13:
            return (col, col)
        elif col == 13:
            return None
        elif 14 <= col <= 33:
            if col > 26:
                if col == 30:
                    return (26, 30)
                return (col - 1, col) if col > 30 else (col, col)
            return (col - 1, col)
        return None
    return (col, col)

def extract_sheet_name_string(sheet_name):
    return CLEAN_NAME_PATTERN.sub('', sheet_name)

def normalize_value(value):
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value
    
    if isinstance(value, float) and value.is_integer():
        value = str(int(value))
    else:
        value = str(value)
    
    value = value.replace("_x000D_", "").replace("\r", "").replace("\n", "")
    value = value.replace('"', '').strip()
    value = TIME_PATTERN.sub(r'\1(', value)
    
    if not value:
        return None
    
    if value in {None, "0", "0:00", "00:00:00", "12:00:00午前"}:
        return None
    
    for pattern in ['%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d']:
        try:
            return datetime.strptime(value, pattern)
        except ValueError:
            continue
    
    try:
        num = float(value)
        return str(int(num)) if num.is_integer() else num
    except ValueError:
        return value

def compare_excel_files(v1_path, result_path):
    try:
        v1_wb = openpyxl.load_workbook(v1_path, read_only=True, data_only=True)
        res_wb = openpyxl.load_workbook(result_path, read_only=True, data_only=True)
        
        visible_sheets_v1 = [(s.title, extract_sheet_name_string(s.title)) 
                             for s in v1_wb.worksheets if s.sheet_state == 'visible']
        visible_sheets_res = [(s.title, extract_sheet_name_string(s.title)) 
                             for s in res_wb.worksheets if s.sheet_state == 'visible']
        
        v1_sheets = {clean: orig for orig, clean in visible_sheets_v1}
        res_sheets = {clean: orig for orig, clean in visible_sheets_res}
        common_sheets = set(v1_sheets.keys()) & set(res_sheets.keys())
        
        if not common_sheets:
            print(f"Error: No common sheets found between {v1_path} and {result_path}")
            return {"ERROR": -1}
        
        results = {}
        for sheet_name in common_sheets:
            v1_ws = v1_wb[v1_sheets[sheet_name]]
            res_ws = res_wb[res_sheets[sheet_name]]
            
            # Precompute columns to compare
            max_col = max(v1_ws.max_column, res_ws.max_column)
            columns = []
            for col in range(1, max_col + 1):
                cols = get_comparison_columns(col, v1_sheets[sheet_name])
                if cols:
                    columns.append((cols[0]-1, cols[1]-1))  # Convert to 0-based
            
            # Read data into memory
            v1_data = [row for row in v1_ws.iter_rows(values_only=True)]
            res_data = [row for row in res_ws.iter_rows(values_only=True)]
            max_rows = max(len(v1_data), len(res_data))
            
            mismatches = []
            for row_idx in range(10,max_rows):
                v1_row = v1_data[row_idx] if row_idx < len(v1_data) else []
                res_row = res_data[row_idx] if row_idx < len(res_data) else []
                
                for v1_col, res_col in columns:
                    v1_val = v1_row[v1_col] if v1_col < len(v1_row) else None
                    res_val = res_row[res_col] if res_col < len(res_row) else None
                    
                    nv1 = normalize_value(v1_val)
                    nres = normalize_value(res_val)
                    if nv1 != nres:
                        mismatches.append({
                            'v1_row': row_idx+1,
                            'v1_col': v1_col+1,
                            'v2_row': row_idx+1,
                            'v2_col': res_col+1,
                            'v1': v1_val,
                            'v2': res_val
                        })
            
            results[res_sheets[sheet_name]] = {
                'yellow_count': 0,  # Styling not supported in read-only mode
                'mismatches': mismatches
            }
        
        v1_wb.close()
        res_wb.close()
        return results
    
    except Exception as e:
        print(f'Error comparing files: {e}')
        return {"ERROR": str(e)}

def process_folder(main_folder):
    results = {}
    for root, dirs, files in os.walk(main_folder):
        if 'V1' in dirs and 'Result' in dirs:
        
            v1_dir = os.path.join(root, 'V1')
            res_dir = os.path.join(root, 'Result')
            
            print(f"Processing folder: {root}")
            print(f"V1 Directory: {v1_dir}")
            print(f"Result Directory: {res_dir}")
            
            for v1_file in os.listdir(v1_dir):
                if not v1_file.endswith(('.xlsx', '.xls')):
                    continue
                base_name = os.path.splitext(v1_file)[0]
                res_files = [f for f in os.listdir(res_dir) if f.startswith(f"X_{base_name}") and f.endswith(('.xlsx', '.xls'))]
                
                if res_files:
                    res_file = res_files[0]
                    v1_path = os.path.join(v1_dir, v1_file)
                    res_path = os.path.join(res_dir, res_file)
                    
                    print(f"Comparing: {v1_file} and {res_file}")
                    try:
                        result = compare_excel_files(v1_path, res_path)
                        key = (v1_path, res_path)
                        results[key] = result
                    except Exception as e:
                        print(f'Error processing {v1_file}: {e}')
    return results

def generate_markdown_report(results):
    report = "## Excel Comparison Report\n\n"
    report += "This report summarizes the comparison of Excel files.\n\n"
    report += "### File Comparison Summary\n\n"

    for (v1_path, res_path), result in results.items():
        v1_filename = os.path.basename(v1_path)
        res_filename = os.path.basename(res_path)
        
        report += f"**Compared Files:**\n\n"
        report += f"* V1 File: {v1_filename}\n\n"
        report += f"* Result File: {res_filename}\n\n"
        
        if "ERROR" in result:
            report += f"**Error:** {result['ERROR']}\n\n"
            continue  # Skip to the next file pair
        
        for sheet_name, sheet_result in result.items():
            report += f"**Sheets Compared:**\n\n"
            report += f"* {sheet_name}\n\n"
            
            if not sheet_result['mismatches']:
                report += "    * No mismatches found.\n\n"
            else:
                report += "    * Mismatches:\n\n"
                for mismatch in sheet_result['mismatches']:
                    report += (
                        f"        * Row {mismatch['v1_row']}, Column {mismatch['v1_col']} (V1) "
                        f"vs Row {mismatch['v2_row']}, Column {mismatch['v2_col']} (Result):\n\n"
                    )
                    report += f"            * V1 Value: {mismatch['v1']}\n\n"
                    report += f"            * Result Value: {mismatch['v2']}\n\n"
    report += "**End of Report**"
    return report

def main():
    root = create_root()
    try:
        folder = select_directory(root, "Select comparison root folder")
        if not folder:
            print("No folder selected. Exiting.")
            return
        results = process_folder(folder)
        if not results:
            print("No files to compare or no mismatches found.")
            return
        report_content = generate_markdown_report(results)
        print(report_content) # Output the report
        # You can also save it to a file:
        with open("comparison_report.md", "w", encoding="utf-8") as f:
            f.write(report_content)
        print("Comparison complete. Report generated.")
    finally:
        root.destroy()

if __name__ == "__main__":
    main()
