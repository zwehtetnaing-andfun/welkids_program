import os
import openpyxl
import sys
import tkinter as tk
import logging
import re
from datetime import datetime
from openpyxl.styles import PatternFill
from tkinter import messagebox, filedialog


# Set up logging configuration
def setup_logging(debug_level):
    # Create logs directory if it doesn't exist
    os.makedirs('logs', exist_ok=True)
    
    # Create a timestamp for the log file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = f'logs/excel_comparison_{timestamp}.log'
    
    # Define the logging format
    log_format = '%(asctime)s - %(levelname)s - %(message)s'
    
    # Set up logging level based on debug_level parameter
    if debug_level == 'DEBUG':
        logging_level = logging.DEBUG
    elif debug_level == 'INFO':
        logging_level = logging.INFO
    else:
        logging_level = logging.WARNING
    
    # Configure logging
    logging.basicConfig(
        level=logging_level,
        format=log_format,
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    logging.info(f'Logging initialized at level: {debug_level}')


def create_root():
    root = tk.Tk()
    root.withdraw()
    return root

def find_timeslot_column(sheet):
    """Find the column containing '外出時間' and return its index."""
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and '外出時間' in cell.value:
                logging.info(f"Found '外出時間' in column {cell.column_letter}")
                return cell.column
    return None


def select_directory(root, prompt):
    logging.debug(f'Showing directory selection dialog with prompt: {prompt}')
    messagebox.showinfo('情報', prompt)
    folder_selected = filedialog.askdirectory(title=prompt)
    if folder_selected:
        logging.info(f'Selected directory: {folder_selected}')
    else:
        logging.warning('No directory selected')
    return folder_selected

def show_message(title, message):
    logging.debug(f'Showing message box - Title: {title}, Message: {message}')
    messagebox.showinfo(title, message)

def normalize_time_format(time_str):
    """
    Normalize time string format by removing leading zeros, seconds, and standardizing separators.
    Examples:
        "08:00:00" -> "8:00"
        "8:00" -> "8:00"
        "09:30:00" -> "9:30"
    """
    if not isinstance(time_str, str):
        time_str = str(time_str)
    
    # Remove any seconds portion if it exists
    if time_str.count(':') == 2:
        time_str = ':'.join(time_str.split(':')[:2])
    
    # Split into hours and minutes
    if ':' in time_str:
        hours, minutes = time_str.split(':')
        # Remove leading zeros from hours
        hours = str(int(hours))
        return f"{hours}:{minutes}"
    
    return time_str

def normalize_time_range_symbols(time_str):
    """Normalize time format by removing variations in symbols and ensuring consistent spacing."""
    return time_str.replace('〜', '~').replace('～', '~').strip()

def format_time_range(time_str):
    """Standardize time range format to ensure consistent comparison."""
    time_str = normalize_time_range_symbols(time_str)  # Use the renamed function
    parts = time_str.split('~')
    if len(parts) == 2:
        start_time = normalize_time_format(parts[0].strip())  # Use new normalization
        end_time = normalize_time_format(parts[1].strip())    # Use new normalization
        return f"{start_time}~{end_time}"
    return time_str

def compare_time_parts(time1, time2):
    """Compare two time strings by removing leading zeros from hours."""
    if ':' not in time1 or ':' not in time2:
        return time1 == time2
    
    hour1, minute1 = time1.split(':')
    hour2, minute2 = time2.split(':')
    
    # Remove leading zeros for comparison
    hour1 = str(int(hour1))
    hour2 = str(int(hour2))
    
    return hour1 == hour2 and minute1 == minute2

def is_datetime_string(value):
    """Check if a string represents a datetime."""
    if not isinstance(value, str):
        return False
    
    # Common datetime patterns in Excel
    patterns = [
        '%Y/%m/%d %H:%M:%S',
        '%Y/%m/%d %H:%M:%S 午前',
        '%Y/%m/%d %H:%M:%S 午後',
        '%Y/%m/%d %H:%M',
        '%Y/%m/%d %H:%M 午前',
        '%Y/%m/%d %H:%M 午後',
        '%Y/%m/%d'
    ]
    
    for pattern in patterns:
        try:
            datetime.strptime(value, pattern)
            return True
        except ValueError:
            continue
    return False

def extract_date_part(value):
    """Extract just the date part from a datetime string or object."""
    if isinstance(value, datetime):
        return value.strftime('%Y/%m/%d')
    
    if not isinstance(value, str):
        return value
        
    # Try to extract date from string
    patterns = [
        '%Y-%m-%d %H:%M:%S',  # Added hyphen format
        '%Y/%m/%d %H:%M:%S',
        '%Y/%m/%d %H:%M:%S 午前',
        '%Y/%m/%d %H:%M:%S 午後',
        '%Y/%m/%d %H:%M',
        '%Y/%m/%d %H:%M 午前',
        '%Y/%m/%d %H:%M 午後',
        '%Y/%m/%d',
        '%Y-%m-%d'  # Added hyphen format without time
    ]
    
    for pattern in patterns:
        try:
            dt = datetime.strptime(value, pattern)
            return dt.strftime('%Y/%m/%d')
        except ValueError:
            continue
            
    # If no pattern matches, try to extract date using regex
    date_pattern = r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})'
    match = re.search(date_pattern, str(value))
    if match:
        year, month, day = match.groups()
        try:
            return f"{year}/{int(month):02d}/{int(day):02d}"
        except ValueError:
            pass
            
    return value

def is_time_string(value):
    """Check if a string represents a time value."""
    if not isinstance(value, str):
        return False
    
    # Normalize the value first
    value = normalize_time_format(value)
    
    # Remove common time separators and spaces
    value = value.replace(':', '').replace('：', '').strip()
    
    # Check if it's a 3 or 4 digit number (e.g., 830 or 1430)
    if value.isdigit() and (len(value) == 3 or len(value) == 4):
        # Extract hours and minutes
        if len(value) == 3:
            hours = int(value[0])
            minutes = int(value[1:])
        else:
            hours = int(value[:2])
            minutes = int(value[2:])
        
        # Validate hours and minutes
        return 0 <= hours <= 23 and 0 <= minutes <= 59
    
    return False

def normalize_time_format(time_str):
    """
    Normalize time string format by removing leading zeros, seconds, and standardizing separators.
    Examples:
        "08:00:00" -> "8:00"
        "8:00" -> "8:00"
        "09:30:00" -> "9:30"
    """
    if not isinstance(time_str, str):
        time_str = str(time_str)
    
    # Remove any seconds portion if it exists
    if time_str.count(':') == 2:
        time_str = ':'.join(time_str.split(':')[:2])
    
    # Split into hours and minutes
    if ':' in time_str:
        hours, minutes = time_str.split(':')
        # Remove leading zeros from hours
        hours = str(int(hours))
        return f"{hours}:{minutes}"
    
    return time_str
        
def get_comparison_columns(col, sheet_name):
    """
    Get the corresponding column numbers for comparison between sheets.
    
    Args:
        col (int): Original column number
        sheet_name (str): Name of the sheet
    
    Returns:
        tuple: (col1, col2) where col1 is for sheet1 and col2 is for sheet2
        None if the column should be skipped
    """
    # For sheets containing "勤務表" in their names
    if "シフト" in sheet_name:
        if col < 39:
            return (col, col)
        else:
            return None
        
def get_mapped_column(original_col, is_sheet1=True):
    """
    Map column numbers according to the specified rules.
    
    Args:
        original_col (int): Original column number
        is_sheet1 (bool): True if mapping for sheet1, False for sheet2
    
    Returns:
        int: Mapped column number
    """
    if is_sheet1:
        # Sheet1: Use original column numbers
        return original_col
    else:
        # Sheet2:
        # If column < 13: Use same column number
        # If column >= 13: Add 1 to column number
        if original_col < 13:
            return original_col
        else:
            return original_col + 1

def extract_sheet_name_string(sheet_name):
    """
    Extract meaningful part of the sheet name by removing leading digits and symbols
    (like dot or underscore) but retaining the rest of the original name structure.
    """
    # Remove leading numbers and symbols (e.g., '1.', '2_', etc.)
    cleaned_name = re.sub(r'^[\d._\-]+', '', sheet_name)
    # Return the remaining name as it is (does not strip anything further)
    return cleaned_name

def compare_time_values(time1, time2):
    """
    Compare two time values accounting for different formats.
    Handles cases like "08:00:00" vs "8:00" as equal.
    """
    # Convert both values to strings and normalize
    time1 = normalize_time_format(str(time1).strip())
    time2 = normalize_time_format(str(time2).strip())
    
    # If either isn't a valid time string, they're not equal
    if not (is_time_string(time1) and is_time_string(time2)):
        return False
    
    # Compare the normalized strings
    return time1 == time2

def compare_excel_files(file1_path, file2_path):
    """Compare two Excel files and return comparison result and modified workbook."""
    logging.info(f'Starting comparison of files:\n  File 1: {file1_path}\n  File 2: {file2_path}')
    """ actual file name should be passed to the function """
    file_name = file1_path.split('/')[-1]

    try:
        # Load the Excel files
        logging.debug('Loading workbooks')
        wb1 = openpyxl.load_workbook(file1_path)
        wb2 = openpyxl.load_workbook(file2_path)
        
        # Initialize variables
        mismatch_found = 0
        fill_pattern_yellow = PatternFill(patternType="solid", fgColor='FFFF00')
        
        # Get visible sheets and their string-only names
        visible_sheets1 = [(sheet.title, extract_sheet_name_string(sheet.title)) 
                          for sheet in wb1.worksheets 
                          if sheet.sheet_state == 'visible']
        
        visible_sheets2 = [(sheet.title, extract_sheet_name_string(sheet.title)) 
                          for sheet in wb2.worksheets 
                          if sheet.sheet_state == 'visible']
        
        # Log visible sheets from both workbooks
        logging.info("Visible sheets in V1:")
        for original, string_only in visible_sheets1:
            logging.info(f"  - Original: {original} -> String only: {string_only}")
            
        logging.info("Visible sheets in V2:")
        for original, string_only in visible_sheets2:
            logging.info(f"  - Original: {original} -> String only: {string_only}")
        
        # Create dictionaries to map string-only names to original names
        sheets1_dict = {string: orig for orig, string in visible_sheets1}
        sheets2_dict = {string: orig for orig, string in visible_sheets2}
        
        # Find matching string-only names
        common_string_names = set(sheets1_dict.keys()) & set(sheets2_dict.keys())
        
        if not common_string_names:
            logging.warning('No matching sheet names found between the workbooks')
            show_message("警告", "両方のExcelファイルに同じ名前のシートが見つかりません。")
            return 'X', wb2
            
        # Compare each matching sheet
        for string_name in common_string_names:
            sheet_name1 = sheets1_dict[string_name]
            sheet_name2 = sheets2_dict[string_name]
            
            logging.info(f'\nComparing sheets: {sheet_name1} <-> {sheet_name2}')
            
            sheet1 = wb1[sheet_name1]
            sheet2 = wb2[sheet_name2]
            
            # Get maximum dimensions for comparison
            row_max = max(sheet1.max_row, sheet2.max_row)
            col_max = max(sheet1.max_column, sheet2.max_column)
            logging.debug(f'Sheet dimensions: {row_max} rows x {col_max} columns')
            
            # Find timeslot column if it exists
            timeslot_col = find_timeslot_column(sheet1)

            # Create dictionary to store Column E values and their row numbers
            wb1_col_e = {}
            wb2_col_e = {}

            # Map Column E values for wb1 (column 5 = E)
            for row in range(6, sheet1.max_row + 1):
                value = normalize_value(sheet1.cell(row, 5).value)
                if value is not None:
                    wb1_col_e[value] = row

            # Compare cells
            for row2 in range(6, row_max + 1):
                try:
                    value = normalize_value(sheet2.cell(row2, 5).value)
                    if value is None:
                        continue

                    if value in wb1_col_e:
                        row1 = wb1_col_e[value]
                        logging.debug(f'Found match: wb1 row {row1} matches wb2 row {row2}')
                        
                        # Remove the matched value from wb1_col_e 
                        del wb1_col_e[value]

                        for col in range(3, col_max + 1):
                            try:
                                # Get comparison columns
                                comparison_cols = get_comparison_columns(col, file_name)
                                
                                if comparison_cols is None:
                                    continue  # Skip this column
                                    
                                col1, col2 = comparison_cols
                                value1 = sheet1.cell(row1, col1).value
                                value2 = sheet2.cell(row2, col2).value

                                logging.debug(f'Comparing cell wb1({row1}, {col1}) with wb2({row2}, {col2})')
                                logging.debug(f'Value1: {value1}, Value2: {value2}')

                                # Normalize values
                                value1 = normalize_value(value1)
                                value2 = normalize_value(value2)

                                # Handle None values
                                if value1 is None and value2 is None:
                                    continue
                                if value1 is None or value2 is None:
                                    sheet2.cell(row2, col2).fill = fill_pattern_yellow
                                    mismatch_found += 1
                                    logging.debug(f'Value mismatch at ({row2}, {col2}): {value1} vs {value2}')
                                    continue

                                # Convert to string and strip whitespace if not datetime object
                                if not isinstance(value1, datetime):
                                    value1 = str(value1).strip()
                                if not isinstance(value2, datetime):
                                    value2 = str(value2).strip()

                                # Check if either value is a datetime
                                is_datetime1 = isinstance(value1, datetime) or is_datetime_string(value1)
                                is_datetime2 = isinstance(value2, datetime) or is_datetime_string(value2)

                                if is_datetime1 or is_datetime2:
                                    date1 = extract_date_part(value1)
                                    date2 = extract_date_part(value2)
                                    
                                    if date1 != date2:
                                        sheet2.cell(row2, col2).fill = fill_pattern_yellow
                                        mismatch_found += 1
                                        logging.debug(f'Date mismatch at ({row2}, {col2}): {date1} vs {date2}')
                                    continue

                                # Check if either value is a time string
                                is_time1 = is_time_string(str(value1))
                                is_time2 = is_time_string(str(value2))

                                if is_time1 or is_time2:
                                    if not compare_time_values(value1, value2):
                                        sheet2.cell(row2, col2).fill = fill_pattern_yellow
                                        mismatch_found += 1
                                        logging.debug(f'Time mismatch at ({row2}, {col2}): {value1} vs {value2}')
                                    continue

                                # Handle time range comparison
                                if any(separator in str(value1) or separator in str(value2) 
                                    for separator in ['〜', '～', '~']):
                                        time1_parts = format_time_range(str(value1)).split('~')
                                        time2_parts = format_time_range(str(value2)).split('~')
                                    
                                        if len(time1_parts) == 2 and len(time2_parts) == 2:
                                            start_match = compare_time_parts(time1_parts[0], time2_parts[0])
                                            end_match = compare_time_parts(time1_parts[1], time2_parts[1])
                                            
                                            if not (start_match and end_match):
                                                sheet2.cell(row2, col2).fill = fill_pattern_yellow
                                                mismatch_found += 1
                                                logging.debug(f'Time range mismatch at ({row2}, {col2}): {value1} vs {value2}')
                                            continue

                                # For all other values, compare as strings
                                if str(value1) != str(value2):
                                    if not is_ignored_mismatch(value1, value2):
                                        sheet2.cell(row2, col2).fill = fill_pattern_yellow
                                        mismatch_found += 1
                                        logging.debug(f'Value mismatch at ({row2}, {col2}): {value1} vs {value2}')
                                    
                            except Exception as e:
                                logging.error(f'Error comparing cell ({row2}, {col2}): {str(e)}')
                                mismatch_found += 1
                                continue
                except Exception as e:
                    logging.error(f'Error processing row {row2}: {str(e)}')
                    continue
        
        # Determine final result
        result = 'X' if mismatch_found > 0 else 'O'
        logging.info(f'Comparison completed. Result: {result} (mismatches: {mismatch_found})')
        return result, wb2
        
    except Exception as e:
        logging.error(f'Error during comparison: {str(e)}', exc_info=True)
        raise

def normalize_value(value):
    """Normalize values to handle numeric equivalence, time formats, blank/None equivalence, and remove special characters."""
    # Handle None values and empty strings
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    
    # If it's already a datetime object, return it
    if isinstance(value, datetime):
        return value
    
    # Convert to string and strip whitespace and quotation marks
    if not isinstance(value, str):
        # Handle float values - convert to int if it's a whole number
        if isinstance(value, float) and value.is_integer():
            value = int(value)
    
    # Convert to string and normalize
    value = str(value)
    
    # Remove special characters and normalize whitespace
    value = value.replace("_x000D_", "")  # Remove Excel carriage return
    value = value.replace("\r", "")       # Remove carriage return
    value = value.replace("\n", "")       # Remove line feed
    value = value.replace('"', '')        # Remove double quotes
    value = value.replace('"', '')        # Remove smart quotes (opening)
    value = value.replace('"', '')        # Remove smart quotes (closing)
    value = ' '.join(value.split())       # Normalize whitespace
    value = value.strip()                 # Remove leading/trailing whitespace
    
    # If after cleaning the string is empty, return None
    if value == "":
        return None
    
    # List of patterns to be normalized to None
    normalize_patterns = [None, "0", "0:00", "00:00:00", "12:00:00午前"]
    
    if value in normalize_patterns:
        return None
        
    # Try parsing as datetime with multiple formats
    datetime_patterns = [
        '%Y-%m-%d %H:%M:%S',
        '%Y/%m/%d %H:%M:%S',
        '%Y-%m-%d',
        '%Y/%m/%d'
    ]
    
    for pattern in datetime_patterns:
        try:
            return datetime.strptime(value, pattern)
        except ValueError:
            continue
    
    # If it's a numeric string (like "1.0" or "1")
    try:
        num = float(value)
        if num.is_integer():
            return str(int(num))
    except ValueError:
        pass
        
    return value

def compare_datetime_values(value1, value2):
    """Compare two datetime values, handling various formats."""
    try:
        # First normalize both values
        norm1 = normalize_value(value1)
        norm2 = normalize_value(value2)
        
        # If both are datetime objects, compare them directly
        if isinstance(norm1, datetime) and isinstance(norm2, datetime):
            return norm1.date() == norm2.date()
            
        # If they're strings, compare the date parts
        date1 = extract_date_part(norm1)
        date2 = extract_date_part(norm2)
        
        return date1 == date2
    except Exception as e:
        logging.error(f"Error comparing datetime values: {value1} vs {value2} - {str(e)}")
        return False

def is_ignored_mismatch(value1, value2):

    """Check if the mismatch between value1 and value2 should be ignored."""
    # List of terms to ignore
    ignored_values = [
        "有給", "振替", "特別", "欠勤", "病気", "介護", "育児", "看護", "公休","その他"
    ]

    # Check if either value1 or value2 is in the ignored values list
    if value1 in ignored_values or value2 in ignored_values:
        return True

    """Check if the mismatch between value1 and value2 should be ignored."""
    same_pairs = [
        ("休み", "シフト時間コード-1"),
        ("長期休", "シフト時間コード-1"),
        ("フリー", "シフト時間コード2147483647")
        # Add other ignored pairs if needed
    ]
    return (value1, value2) in same_pairs or (value2, value1) in same_pairs

def process_folder(recompare_folder):
    """Process all subfolders in recompare directory"""
    try:
        # Get all subfolders in recompare directory
        subfolders = [f for f in os.listdir(recompare_folder) 
                     if os.path.isdir(os.path.join(recompare_folder, f))]
        
        logging.info(f'Found {len(subfolders)} subfolders to process')
        
        for subfolder in subfolders:
            subfolder_path = os.path.join(recompare_folder, subfolder)
            v1_path = os.path.join(subfolder_path, 'V1')
            v2_path = os.path.join(subfolder_path, 'V2')
            result_path = os.path.join(subfolder_path, 'result')
            
            # Check if required folders exist
            if not all(os.path.exists(p) for p in [v1_path, v2_path]):
                logging.warning(f'Skipping {subfolder}: V1 or V2 folder missing')
                continue
                
            # Create result folder if it doesn't exist
            os.makedirs(result_path, exist_ok=True)
            logging.info(f'Processing subfolder: {subfolder}')
            
            # Get Excel files from V1
            files_v1 = [f for f in os.listdir(v1_path) 
                       if f.endswith(('.xlsx', '.xls'))]
            
            # Process each file
            for file_name in files_v1:
                base_name = os.path.splitext(file_name)[0]
                file_v2 = f"{base_name}.xlsx"  # Assuming same name in V2
                
                v1_file_path = os.path.join(v1_path, file_name)
                v2_file_path = os.path.join(v2_path, file_v2)
                
                if not os.path.exists(v2_file_path):
                    logging.warning(f'No matching V2 file for {file_name}')
                    continue
                    
                logging.info(f'\nComparing:\nV1: {file_name}\nV2: {file_v2}')
                
                try:
                    # Compare files and get result
                    result, modified_wb = compare_excel_files(v1_file_path, v2_file_path)
                    
                    # Save result
                    output_filename = f"{result}_{base_name}.xlsx"
                    output_path = os.path.join(result_path, output_filename)
                    
                    logging.info(f'Saving result to: {output_path}')
                    modified_wb.save(output_path)
                    
                except Exception as e:
                    logging.error(f'Error processing {file_name}: {str(e)}')
                    continue
                    
        return True
        
    except Exception as e:
        logging.error(f'Error in process_folder: {str(e)}', exc_info=True)
        return False

def main():
    # Initialize logging
    setup_logging('DEBUG')
    logging.info('Starting Excel comparison program')
    
    try:
        # Create root window
        root = create_root()
        
        # Select recompare folder
        recompare_folder = select_directory(root, "recompareフォルダーを選択してください")
        if not recompare_folder:
            logging.warning('Folder selection cancelled')
            show_message("フォルダー選択", "フォルダーが選択されていません。終了します...")
            return
            
        logging.info(f'Selected recompare folder: {recompare_folder}')
        show_message("比較を開始します", "比較プロセスを開始しています....")
        
        # Process all subfolders
        success = process_folder(recompare_folder)
        
        if success:
            logging.info('Comparison process completed successfully')
            show_message("比較が完了しました", "比較プロセスが完了しました.")
        else:
            show_message("エラー", "処理中にエラーが発生しました。ログを確認してください。")
            
    except Exception as e:
        logging.error(f'Unexpected error: {str(e)}', exc_info=True)
        messagebox.showerror("エラーが発生しました", f"予期しないエラーが発生しました: {str(e)}")
    finally:
        logging.info('Program finished')
        if root:
            root.destroy()

if __name__ == "__main__":
    main()