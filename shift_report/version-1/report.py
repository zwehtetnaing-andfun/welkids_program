import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import PatternFill
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed

def count_colored_cells_in_sheet(sheet):
    yellow_cells = 0
    fill_pattern_yellow = PatternFill(patternType="solid", fgColor='FFFF00')
    
    try:
        columns_max = min(sheet.max_column, 38)  # Limit to column Z
        rows_max = sheet.max_row
        
        for row in sheet.iter_rows(min_row=5, max_row=rows_max, min_col=3, max_col=columns_max):
            for cell in row:
                if cell.fill == fill_pattern_yellow:
                    yellow_cells += 1
        return yellow_cells
    except Exception as e:
        print(f"Error processing sheet: {str(e)}")
        return 0

def process_excel_file(file_path):
    result = {}
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            colored_count = count_colored_cells_in_sheet(ws)
            result[sheet_name] = colored_count
        wb.close()
    except Exception as e:
        print(f"Error processing sheet: {str(e)}")
        result["ERROR"] = -1
    return result

def select_folder():
    root = tk.Tk()
    root.withdraw()
    folder_selected = filedialog.askdirectory(title="Select the source folder")
    root.destroy()
    return folder_selected

def find_result_excels(base_folder):
    base_path = Path(base_folder)
    return [str(p) for p in base_path.rglob("Result/*.xlsx") if not p.name.startswith("~$")]

def generate_markdown_report(report_data, output_path=None):
    if output_path is None:
        output_path = datetime.now().strftime("report_%Y%m%d_%H%M%S.md")

    school_data = {}
    for file_path, sheet_data in report_data.items():
        school_id = os.path.basename(os.path.dirname(os.path.dirname(file_path)))
        school_data.setdefault(school_id, []).append((file_path, sheet_data))

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("# Excel Compare Report\n\n")
        for school_id, files in school_data.items():
            f.write(f"school id = {school_id}\n{'-' * 80}\n")
            f.write("| workbook | sheet | X | |\n| --- | --- | --- | --- |\n")
            for file_path, sheet_data in files:
                file_name = os.path.basename(file_path)
                for i, (sheet, count) in enumerate(sheet_data.items()):
                    warning = "❌" if count < 0 else "⚠️" if count > 0 else ""
                    f.write(f"| {file_name if i == 0 else ''} | {sheet} | {count} | {warning} |\n")
            f.write("-" * 80 + "\n")

        f.write("\nfinal result report :\n" + "-" * 25 + "\n")
        f.write("| School ID | Status |\n| --- | --- |\n")
        for school_id, files in school_data.items():
            has_warning = any(count > 0 for _, sheet_data in files for count in sheet_data.values())
            status = "⚠️" if has_warning else "✅"
            f.write(f"| {school_id} | {status} |\n")
        f.write("-" * 25)

    print(f"\n✅ Report generated at: {output_path}")

def main():
    folder = select_folder()
    if not folder:
        print("❌ No folder selected.")
        return

    excel_files = find_result_excels(folder)
    if not excel_files:
        print("⚠️ No Excel files found in 'Result' folders.")
        return

    report_data = {}
    with ProcessPoolExecutor(max_workers=4) as executor:
        future_to_file = {executor.submit(process_excel_file, file_path): file_path for file_path in excel_files}
        for future in as_completed(future_to_file):
            file_path = future_to_file[future]
            relative_path = os.path.relpath(file_path, folder)
            print(f"Processed: {relative_path}")
            try:
                report_data[relative_path] = future.result()
            except Exception as e:
                print(f"Error processing {relative_path}: {str(e)}")

    generate_markdown_report(report_data)

if __name__ == "__main__":
    main()